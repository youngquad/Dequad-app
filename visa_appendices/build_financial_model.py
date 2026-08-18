"""DEQUAD financial model — self-funded 3-year edition.

Starting reality:
  - Two co-founders, £3,000 EACH (£6,000 pooled) in the bank, MVP already shipped.
  - On the NatWest Accelerator (London cohort) — joined 16 March 2026. Office
    co-working, legal advice and accountancy support all provided FREE
    through the programme for the first 12 months (Y1 only).
  - Previously completed the Santander Universities Pre-Incubator programme
    (2025) — another independent validation point.
  - Prospective anchor pilot: University of Bedfordshire (lead founder served
    as SU President 2021-2023, direct senior-management relationship) — this
    is an EARLY, INFORMAL CONVERSATION ONLY. No agreement, LOI or date is
    signed, and nothing is guaranteed.

Key change from earlier drafts of this model: NO pre-seed, seed, Series A,
grant or R&D tax credit is assumed anywhere. The £6,000 founder capital plus
revenue generated in the period is the ONLY funding in this plan. Founders
draw no salary until Year 3, and only then to the extent revenue supports it
— not gated on any funding round, because none is assumed.

This produces a model that is credible for an Innovator visa endorsement:
  - Cash never goes negative in any month of Year 1 or any year of the
    3-year forecast — on founder capital and revenue alone.
  - No institutional revenue is assumed in Y1 (nothing is signed).
  - Growth (the plan's one hire, modest founder pay) is triggered by
    revenue actually landing, not by an external funding event.

Run with:  python build_financial_model.py   (run from visa_appendices/)
Output:    DEQUAD_Financial_Model.xlsx (written next to this script)
"""
from __future__ import annotations

from pathlib import Path

from openpyxl import Workbook
from openpyxl.styles import Alignment, Border, Font, PatternFill, Side
from openpyxl.utils import get_column_letter

OUT = str(Path(__file__).resolve().parent / "DEQUAD_Financial_Model.xlsx")

NAVY, BLUE, SOFT, WHITE, GREY, AMBER = "0F2942", "5B9BD5", "EDF4FB", "FFFFFF", "F1F5F9", "FEF3C7"
thin = Side(border_style="thin", color="DDE8F2")
border_all = Border(left=thin, right=thin, top=thin, bottom=thin)
header_fill = PatternFill("solid", fgColor=NAVY)
sub_fill = PatternFill("solid", fgColor=SOFT)
total_fill = PatternFill("solid", fgColor=GREY)
amber_fill = PatternFill("solid", fgColor=AMBER)

bold_white = Font(bold=True, color=WHITE, name="Calibri", size=11)
bold = Font(bold=True, name="Calibri", size=11)
regular = Font(name="Calibri", size=11)
italic_small = Font(name="Calibri", size=9, italic=True, color="4F6076")
title_font = Font(bold=True, size=14, color=NAVY, name="Calibri")

center = Alignment(horizontal="center", vertical="center")
right_a = Alignment(horizontal="right", vertical="center")
left_a = Alignment(horizontal="left", vertical="center", wrap_text=True)


def gbp(n): return '£#,##0;[Red](£#,##0);"-"'


def write_header(ws, title, cols, col_widths=None):
    ws["A1"] = title
    ws["A1"].font = title_font
    ws["A2"] = ("All figures in GBP, net of VAT. Forecast — not a guarantee. "
                "No institutional revenue is assumed in Y1; no pre-seed, seed or other "
                "external funding is assumed anywhere in this model. NatWest Accelerator "
                "covers office, legal and accountancy in-kind for Year 1 only.")
    ws["A2"].font = italic_small
    ws.merge_cells(start_row=2, start_column=1, end_row=2, end_column=1 + len(cols))

    row = 4
    for j, h in enumerate(cols, start=2):
        c = ws.cell(row=row, column=j, value=h)
        c.fill = header_fill
        c.font = bold_white
        c.alignment = center
        c.border = border_all
    if col_widths:
        for i, w in enumerate(col_widths, start=1):
            ws.column_dimensions[get_column_letter(i)].width = w


def write_row(ws, row, label, values, *, bold_row=False, sub=False, total=False, money=True):
    c = ws.cell(row=row, column=1, value=label)
    c.alignment = left_a
    c.font = bold if (bold_row or total) else regular
    c.border = border_all
    if sub: c.fill = sub_fill
    if total: c.fill = total_fill
    for j, v in enumerate(values, start=2):
        cell = ws.cell(row=row, column=j, value=v)
        cell.alignment = right_a
        cell.font = bold if (bold_row or total) else regular
        cell.border = border_all
        if money and isinstance(v, (int, float)):
            cell.number_format = gbp(v)
        if sub: cell.fill = sub_fill
        if total: cell.fill = total_fill


def sub_row(ws, row, label, ncols):
    cell = ws.cell(row=row, column=1, value=label)
    cell.font, cell.fill, cell.alignment, cell.border = bold, sub_fill, left_a, border_all
    for j in range(2, ncols + 2):
        ws.cell(row=row, column=j).fill = sub_fill
        ws.cell(row=row, column=j).border = border_all


# ============================================================
# README
# ============================================================
def sheet_readme(wb):
    ws = wb.active
    ws.title = "README"
    ws.column_dimensions["A"].width = 110
    ws["A1"] = "DEQUAD — Financial Model (Self-Funded 3-Year Edition)"
    ws["A1"].font = Font(bold=True, size=16, color=NAVY)
    intro = [
        "",
        "Prepared for: UKES (endorsing body) — UK Innovator Founder Visa",
        "Entity:       DEQUAD Ltd (company in formation, England & Wales)",
        "Founders:     Two co-founders — Yusuf Quadri (CEO) and Yusuff Adeagbo (CTO)",
        "Starting cash: £6,000 (£3,000 from each founder, pooled in the company bank) —",
        "               the ONLY funding assumed anywhere in this model.",
        "MVP status:    Production-ready and deployed at https://dequad.co.uk",
        "Accelerator:   NatWest Accelerator (London) — joined 16 March 2026.",
        "               Office, legal and accountancy support provided in-kind for Year 1 only.",
        "Prior accel:   Santander Universities Pre-Incubator — completed 2025.",
        "Pilot status:  Early, informal conversation with University of Bedfordshire (lead",
        "               founder = former SU President 2021-2023). NO agreement, LOI or date",
        "               is signed. Modelled as a Y2 upside only — never assumed in Y1.",
        "Currency:      GBP, net of VAT",
        "Horizon:       3 years annual + Year-1 monthly cash flow",
        "",
        "KEY MODEL ASSUMPTIONS (deliberately conservative and self-funded)",
        "  - NO pre-seed, seed, Series A, grant or R&D tax credit is assumed anywhere.",
        "  - Y1 institutional revenue is £0 — nothing is signed. Premium (B2C) revenue",
        "    is modelled independently of any pilot, from the existing 80-person",
        "    Bedfordshire beta cohort plus modest organic growth.",
        "  - Founders take NO salary in Y1 or Y2 (living on personal savings/freelance",
        "    income). Modest £500/month each begins in Y3, funded entirely by revenue.",
        "  - The plan's only funded hire (Safeguarding & Trust Lead, part-time) starts",
        "    in Y3, contingent on 2+ paying universities, funded entirely by revenue.",
        "  - Office, legal and accountancy carry a nominal £0 cash cost in Y1 because",
        "    those services are provided FREE by the NatWest Accelerator (Y1 only).",
        "  - Marketing is founder-led and low-cost throughout — no paid-acquisition",
        "    budget in Y1; spend grows only in line with actual revenue in Y2-Y3.",
        "  - Y1 starts with the existing MVP, so no upfront engineering capex.",
        "",
        "WORKBOOK STRUCTURE",
        "  1. P&L 3yr            — Annual Profit & Loss",
        "  2. Cash Flow 3yr      — Annual cash flow forecast",
        "  3. Cash Flow Y1 (mo)  — Monthly cash flow for Year 1",
        "  4. Balance Sheet 3yr  — Year-end balance sheets",
        "  5. Revenue W1         — Revenue detail by service",
        "  6. Cost of Sales W2   — Direct cost detail",
        "  7. Payroll W3         — Salaries, NI and pension",
        "  8. Marketing W4       — Marketing spend by channel",
        "  9. R&D W5             — R&D activity (unpaid founder time; no tax credit assumed)",
        " 10. Fixed Assets       — CAPEX schedule",
        " 11. Accelerator Value  — Quantified in-kind value of NatWest support (Y1 only)",
        " 12. Startup Loan       — Empty (no debt taken; no external funding assumed at all)",
        "",
        "All assumptions map to DEQUAD_UKES_Business_Plan.md (v5.0, self-funded revision).",
        "Yellow cells indicate user-editable inputs.",
    ]
    for i, line in enumerate(intro, start=2):
        c = ws.cell(row=i, column=1, value=line)
        c.alignment = Alignment(wrap_text=True, vertical="top")
        if line.startswith(("KEY MODEL", "WORKBOOK STRUCTURE")):
            c.font = Font(bold=True, color=NAVY)


# ============================================================
# Profit & Loss — self-funded, conservative numbers
# ============================================================
PL = {
    # Revenue
    "rev_uni":     (0, 10_000, 30_000),      # 0 / 0.5 / 1.5 avg paying institutions — nothing signed in Y1
    "rev_premium": (600, 6_000, 18_000),     # 20 / 100 / 300 avg paying subscribers
    "rev_icb":     (0, 0, 0),                # no NHS ICB revenue assumed in this 3-year plan

    # COGS
    "cogs_cloud":  (600, 1_200, 2_400),
    "cogs_llm":    (400, 900, 1_800),
    "cogs_stripe": (50, 90, 180),
    "cogs_msg":    (150, 410, 800),
    "cogs_tools":  (200, 500, 920),

    # Overheads
    "salaries":    (0, 0, 20_000),           # 2 founders @ £500/mo + 1 part-time hire, Y3 only
    "ni":          (0, 0, 900),
    "pension":     (0, 0, 350),
    "benefits":    (0, 0, 250),
    "software":    (600, 1_200, 1_800),
    "office":      (0, 300, 600),
    "legal_acc":   (300, 600, 1_200),
    "marketing":   (900, 2_500, 5_000),
    "insurance":   (480, 700, 900),
    "misc":        (400, 700, 1_000),
}

REV_Y = [PL["rev_uni"][i] + PL["rev_premium"][i] + PL["rev_icb"][i] for i in range(3)]
COGS_Y = [sum(PL[k][i] for k in ("cogs_cloud", "cogs_llm", "cogs_stripe", "cogs_msg", "cogs_tools")) for i in range(3)]
GP_Y = [REV_Y[i] - COGS_Y[i] for i in range(3)]
OH_Y = [sum(PL[k][i] for k in ("salaries", "ni", "pension", "benefits", "software",
                               "office", "legal_acc", "marketing", "insurance", "misc"))
        for i in range(3)]
EBITDA_Y = [GP_Y[i] - OH_Y[i] for i in range(3)]
DEPR_Y = [-150, -400, -700]
PBT_Y = [EBITDA_Y[i] + DEPR_Y[i] for i in range(3)]
TAX_Y = [0, -500, -1_200]
PAT_Y = [PBT_Y[i] + TAX_Y[i] for i in range(3)]


def sheet_pl(wb):
    ws = wb.create_sheet("P&L 3yr")
    write_header(ws, "Annual Profit & Loss Forecast", ["Year 1", "Year 2", "Year 3", "TOTAL"],
                 [44, 16, 16, 16, 16])

    blocks = [
        ("Revenue", [
            ("Service 1 — University SaaS subscription (contingent, not signed)", PL["rev_uni"]),
            ("Service 2 — DEQUAD Premium (B2C)", PL["rev_premium"]),
            ("Service 3 — NHS ICB (not assumed in this 3-year plan)", PL["rev_icb"]),
        ], REV_Y, "Total Revenue"),

        ("Cost of Sales (W2)", [
            ("Cloud hosting (Render, Cloudflare)", PL["cogs_cloud"]),
            ("LLM / safeguarding inference", PL["cogs_llm"]),
            ("Stripe processing", PL["cogs_stripe"]),
            ("SMS & email (Twilio, SendGrid)", PL["cogs_msg"]),
            ("Customer-success tooling", PL["cogs_tools"]),
        ], COGS_Y, "Total Cost of Sales"),
    ]

    r = 5
    for sub, items, total_v, total_lbl in blocks:
        sub_row(ws, r, sub, 4); r += 1
        for label, vals in items:
            write_row(ws, r, label, list(vals) + [sum(vals)]); r += 1
        write_row(ws, r, total_lbl, list(total_v) + [sum(total_v)], total=True); r += 1
        r += 1

    write_row(ws, r, "Gross Profit / (Loss)", GP_Y + [sum(GP_Y)], total=True); r += 1
    write_row(ws, r, "Gross margin %",
              [f"{round(GP_Y[i] / REV_Y[i] * 100, 1)}%" if REV_Y[i] else "n/a" for i in range(3)] + [""],
              money=False); r += 2

    sub_row(ws, r, "Overhead Expenditure", 4); r += 1
    for label, key in [
        ("Salaries (W3) — £0 until Y3", "salaries"),
        ("Employer NI", "ni"),
        ("Employer pension (3%)", "pension"),
        ("Other employment costs (kit, training)", "benefits"),
        ("Software subscriptions", "software"),
        ("Office (£0 in Y1 — NatWest Accelerator)", "office"),
        ("Legal & accountancy (out-of-programme only)", "legal_acc"),
        ("Marketing (W4)", "marketing"),
        ("Insurance", "insurance"),
        ("Business support / miscellaneous", "misc"),
    ]:
        write_row(ws, r, label, list(PL[key]) + [sum(PL[key])]); r += 1
    write_row(ws, r, "Total Overhead Expenditure", OH_Y + [sum(OH_Y)], total=True); r += 2

    write_row(ws, r, "EBITDA", EBITDA_Y + [sum(EBITDA_Y)], total=True); r += 1
    write_row(ws, r, "EBITDA margin %",
              [f"{round(EBITDA_Y[i] / REV_Y[i] * 100, 1)}%" if REV_Y[i] else "n/a" for i in range(3)] + [""],
              money=False); r += 1
    write_row(ws, r, "Depreciation & amortisation", DEPR_Y + [sum(DEPR_Y)]); r += 1
    write_row(ws, r, "Operating profit / (loss)", PBT_Y + [sum(PBT_Y)], total=True); r += 1
    write_row(ws, r, "Corporation Tax (paid in arrears — see Cash Flow)", TAX_Y + [sum(TAX_Y)]); r += 1
    write_row(ws, r, "Profit / (loss) after tax", PAT_Y + [sum(PAT_Y)], total=True)


# ============================================================
# Cash Flow 3yr — self-funded only
# ============================================================
RECEIPTS = {
    "sales": REV_Y,
}
EXPENSES = {
    "cogs":      [-c for c in COGS_Y],
    "payroll":   [-(PL["salaries"][i] + PL["ni"][i] + PL["pension"][i] + PL["benefits"][i]) for i in range(3)],
    "marketing": [-m for m in PL["marketing"]],
    "software":  [-s for s in PL["software"]],
    "office":    [-o for o in PL["office"]],
    "legal_acc": [-l for l in PL["legal_acc"]],
    "insurance": [-i for i in PL["insurance"]],
    "misc":      [-m for m in PL["misc"]],
    "capex":     (-600, -600, -1_200),
    "tax_paid":  (0, 0, -500),   # Y2's accrued tax liability, paid in arrears during Y3
}
TOTAL_R = [sum(v[i] for v in RECEIPTS.values()) for i in range(3)]
TOTAL_E = [sum(v[i] for v in EXPENSES.values()) for i in range(3)]
SURPLUS = [TOTAL_R[i] + TOTAL_E[i] for i in range(3)]

# £6,000 founder capital is the Year-1 OPENING balance (already invested Day 1),
# not an in-year receipt — matching the Business Plan's presentation (Section 12.1).
CASH_OPEN = [6_000]
for s in SURPLUS[:-1]:
    CASH_OPEN.append(CASH_OPEN[-1] + s)
CASH_CLOSE = [CASH_OPEN[i] + SURPLUS[i] for i in range(3)]


def sheet_cf(wb):
    ws = wb.create_sheet("Cash Flow 3yr")
    write_header(ws, "Annual Cash Flow Forecast (Self-Funded)", ["Year 1", "Year 2", "Year 3"], [46, 16, 16, 16])

    r = 5
    write_row(ws, r, "Opening cash balance", list(CASH_OPEN)); r += 1
    write_row(ws, r, "  (Year 1 opening = £6,000 founder equity, already invested Day 1)", ["", "", ""],
              money=False); r += 2

    sub_row(ws, r, "RECEIPTS", 3); r += 1
    write_row(ws, r, "Cash from sales (collected)", list(RECEIPTS["sales"])); r += 1
    write_row(ws, r, "Total Receipts", TOTAL_R, total=True); r += 2

    sub_row(ws, r, "EXPENDITURE", 3); r += 1
    for label, key in [
        ("Cost of sales", "cogs"),
        ("Payroll (gross + NI + pension + benefits)", "payroll"),
        ("Marketing", "marketing"),
        ("Software subscriptions", "software"),
        ("Office (£0 Y1 — NatWest Accelerator)", "office"),
        ("Legal & accountancy (out-of-programme)", "legal_acc"),
        ("Insurance", "insurance"),
        ("Business support / misc", "misc"),
        ("Fixed assets capex", "capex"),
        ("Corporation tax (prior-year liability, paid in arrears)", "tax_paid"),
    ]:
        write_row(ws, r, label, list(EXPENSES[key])); r += 1
    write_row(ws, r, "Total Expenditure", TOTAL_E, total=True); r += 2

    write_row(ws, r, "Cash surplus / (deficit)", SURPLUS, total=True); r += 1
    write_row(ws, r, "Closing cash balance", list(CASH_CLOSE), total=True); r += 2

    ws.cell(row=r, column=1,
            value="No pre-seed, seed, Series A, grant or R&D tax credit is assumed anywhere in this table. "
                  "Closing cash stays positive every year, funded entirely by the £6,000 founder capital plus "
                  "revenue — the plan does not depend on any pilot converting or any investor closing.").font = italic_small
    ws.merge_cells(start_row=r, start_column=1, end_row=r, end_column=4)


# ============================================================
# Year 1 monthly cash flow — critical for UKES
# ============================================================
def sheet_cf_monthly(wb):
    ws = wb.create_sheet("Cash Flow Y1 (mo)")
    months = ["M1\nJun", "M2\nJul", "M3\nAug", "M4\nSep", "M5\nOct", "M6\nNov",
              "M7\nDec", "M8\nJan", "M9\nFeb", "M10\nMar", "M11\nApr", "M12\nMay"]
    cols = months + ["Y1 total"]
    write_header(ws, "Year 1 Monthly Cash Flow Forecast (Self-Funded)", cols, [42] + [11] * 13)

    # No institutional revenue and no external funding at any point in Y1.
    # Premium (B2C) revenue begins M7 when Stripe billing goes live for the
    # existing 80-person Bedfordshire beta cohort — independent of whether
    # any pilot conversation converts.
    inflow_sale = [0, 0, 0, 0, 0, 0, 50, 70, 90, 110, 130, 150]
    assert sum(inflow_sale) == PL["rev_premium"][0] == 600
    total_in = inflow_sale

    # Outflows — self-funded, no payroll in Y1 (founders unpaid throughout)
    out_cogs = [60, 60, 70, 80, 90, 100, 110, 120, 130, 140, 150, 290]
    out_sw = [50] * 12
    out_legal = [300, 0, 0, 0, 0, 0, 0, 0, 0, 0, 0, 0]   # incorporation, IP filings
    out_ins = [0, 0, 48, 48, 48, 48, 48, 48, 48, 48, 48, 48]
    out_mkt = [0, 0, 50, 80, 80, 80, 90, 90, 90, 90, 90, 160]
    out_misc = [125, 25, 25, 25, 25, 25, 25, 25, 25, 25, 25, 25]
    out_capex = [500, 0, 0, 0, 0, 0, 0, 0, 0, 0, 0, 100]   # laptops M1; minor addition M12

    assert sum(out_cogs) == COGS_Y[0] == 1_400
    assert sum(out_sw) == PL["software"][0] == 600
    assert sum(out_legal) == PL["legal_acc"][0] == 300
    assert sum(out_ins) == PL["insurance"][0] == 480
    assert sum(out_mkt) == PL["marketing"][0] == 900
    assert sum(out_misc) == PL["misc"][0] == 400
    assert sum(out_capex) == 600

    total_out = [-(a + b + c + d + e + f + g) for a, b, c, d, e, f, g in
                 zip(out_cogs, out_sw, out_legal, out_ins, out_mkt, out_misc, out_capex)]
    assert sum(total_out) == -4_680

    # Cash balance walk — opens at £6,000 (founder capital, Day 1)
    bal = [6_000] + [0] * 12
    for i in range(12):
        bal[i + 1] = bal[i] + total_in[i] + total_out[i]
    opening = bal[:-1]
    closing = bal[1:]
    assert round(closing[-1]) == 1_920

    rows = [
        ("Opening cash balance", opening, False),
        ("RECEIPTS", None, "sub"),
        ("  Cash from premium (B2C) sales — Stripe live from M7", inflow_sale, False),
        ("Total Receipts", total_in, True),
        ("EXPENDITURE", None, "sub"),
        ("  Cost of sales", [-x for x in out_cogs], False),
        ("  Software subscriptions", [-x for x in out_sw], False),
        ("  Legal & accountancy (incorporation, IP filings)", [-x for x in out_legal], False),
        ("  Insurance", [-x for x in out_ins], False),
        ("  Marketing (founder-led, low-cost)", [-x for x in out_mkt], False),
        ("  Business support / misc", [-x for x in out_misc], False),
        ("  Fixed assets capex", [-x for x in out_capex], False),
        ("Total Expenditure", total_out, True),
        ("Closing cash balance", closing, True),
    ]
    r = 5
    for label, monthly, style in rows:
        if monthly is None and style == "sub":
            sub_row(ws, r, label, 13); r += 1; continue
        values = list(monthly) + [sum(monthly)]
        write_row(ws, r, label, values, total=(style is True))
        if label.startswith("Closing"):
            for i, v in enumerate(monthly, start=2):
                if isinstance(v, (int, float)) and v < 0:
                    ws.cell(row=r, column=i).fill = amber_fill
        r += 1
    ws.cell(row=r + 1, column=1,
            value="No founder salary is drawn in Year 1, and no institutional revenue or external funding is "
                  "assumed. Cash declines steadily from the £6,000 opening balance as the business absorbs "
                  "incorporation and running costs, then stabilises once modest premium-subscription revenue "
                  "begins in M7 — closing the year at roughly £1,920, positive throughout.").font = italic_small
    ws.merge_cells(start_row=r + 1, start_column=1, end_row=r + 1, end_column=14)


# ============================================================
# Balance Sheet
# ============================================================
def sheet_bs(wb):
    ws = wb.create_sheet("Balance Sheet 3yr")
    write_header(ws, "Annual Balance Sheet Forecast", ["Year 1", "Year 2", "Year 3"], [44, 16, 16, 16])

    # CAPEX (cumulative) — small because MVP already built
    capex = [-x for x in EXPENSES["capex"]]
    cum_capex = [capex[0], capex[0] + capex[1], capex[0] + capex[1] + capex[2]]
    cum_dep = [-DEPR_Y[0], -(DEPR_Y[0] + DEPR_Y[1]), -(DEPR_Y[0] + DEPR_Y[1] + DEPR_Y[2])]
    nbv = [cum_capex[i] - cum_dep[i] for i in range(3)]

    cash = CASH_CLOSE
    debtors = [0, 500, 1_500]
    payables = [0, 1_000, 2_700]

    # No external investment round — no share premium in any year.
    share_premium = [0, 0, 0]
    share_cap = [6_000, 6_000, 6_000]
    retained = [PAT_Y[0], PAT_Y[0] + PAT_Y[1], PAT_Y[0] + PAT_Y[1] + PAT_Y[2]]
    sh_funds = [share_cap[i] + share_premium[i] + retained[i] for i in range(3)]

    total_assets = [nbv[i] + cash[i] + debtors[i] for i in range(3)]
    total_liab = list(payables)
    net_assets = [total_assets[i] - total_liab[i] for i in range(3)]
    check = [net_assets[i] - sh_funds[i] for i in range(3)]

    rows = [
        ("NON-CURRENT ASSETS", None, "sub"),
        ("Fixed assets (net book value)", nbv, False),
        ("Total non-current assets", nbv, True),
        ("", None, False),
        ("CURRENT ASSETS", None, "sub"),
        ("Cash at bank", cash, False),
        ("Trade receivables (debtors)", debtors, False),
        ("Stock / inventory", [0, 0, 0], False),
        ("Total current assets", [cash[i] + debtors[i] for i in range(3)], True),
        ("", None, False),
        ("TOTAL ASSETS", total_assets, True),
        ("", None, False),
        ("CURRENT LIABILITIES", None, "sub"),
        ("Trade payables / accrued costs", payables, False),
        ("Director's loan account", [0, 0, 0], False),
        ("Total liabilities", total_liab, True),
        ("", None, False),
        ("NET ASSETS", net_assets, True),
        ("", None, False),
        ("CAPITAL & RESERVES", None, "sub"),
        ("Called-up share capital", share_cap, False),
        ("Share premium (none — no external investment round)", share_premium, False),
        ("Profit & loss reserve", retained, False),
        ("Shareholders' funds", sh_funds, True),
        ("", None, False),
        ("Balance check (should be £0)", check, False),
    ]
    r = 5
    for label, vals, style in rows:
        if vals is None:
            if style == "sub":
                sub_row(ws, r, label, 3)
            r += 1; continue
        write_row(ws, r, label, vals, total=(style is True))
        r += 1


# ============================================================
# Revenue W1
# ============================================================
def sheet_w1(wb):
    ws = wb.create_sheet("Revenue W1")
    write_header(ws, "Revenue Detail (W1)", ["Year 1", "Year 2", "Year 3", "TOTAL"],
                 [50, 14, 14, 14, 14])
    rows = [
        ("Total Revenue", REV_Y, "total"),
        ("", None, False),
        ("Service 1 — University SaaS subscription", None, "sub"),
        ("Paying institutions (avg during year) — contingent, not signed", [0, 0.5, 1.5], False),
        ("Price per enrolled student (£/yr)", [2, 2, 2], False),
        ("Average enrolled students per institution", [10_000, 10_000, 10_000], False),
        ("Average contract value (£/yr)", [20_000, 20_000, 20_000], False),
        ("Subtotal revenue", PL["rev_uni"], "total"),
        ("", None, False),
        ("Service 2 — DEQUAD Premium (B2C)", None, "sub"),
        ("Paying student subs (avg, year) — independent of any pilot", [20, 100, 300], False),
        ("Price per student (£/month)", [4.99, 4.99, 4.99], False),
        ("Subtotal revenue", PL["rev_premium"], "total"),
        ("", None, False),
        ("Service 3 — NHS ICB", None, "sub"),
        ("Number of contracts (not assumed in this 3-year plan)", [0, 0, 0], False),
        ("Subtotal revenue", PL["rev_icb"], "total"),
    ]
    r = 5
    for label, vals, style in rows:
        if vals is None:
            if style == "sub":
                sub_row(ws, r, label, 4)
            r += 1; continue
        total_val = sum(vals) if all(isinstance(v, (int, float)) for v in vals) else ""
        write_row(ws, r, label, list(vals) + [total_val], total=(style == "total"),
                  money=isinstance(vals[0], (int, float)))
        r += 1


# ============================================================
# Cost of Sales W2
# ============================================================
def sheet_w2(wb):
    ws = wb.create_sheet("Cost of Sales W2")
    write_header(ws, "Cost of Sales Detail (W2)", ["Year 1", "Year 2", "Year 3", "TOTAL"],
                 [44, 14, 14, 14, 14])
    rows = [
        ("Total Cost of Sales", COGS_Y, "total"),
        ("Cloud hosting (Render, Cloudflare CDN)", PL["cogs_cloud"], False),
        ("LLM / safeguarding inference (OpenAI/Anthropic)", PL["cogs_llm"], False),
        ("Stripe processing", PL["cogs_stripe"], False),
        ("SMS & email (Twilio, SendGrid)", PL["cogs_msg"], False),
        ("Customer-success tooling (Intercom, Hotjar)", PL["cogs_tools"], False),
    ]
    r = 5
    for label, vals, style in rows:
        write_row(ws, r, label, list(vals) + [sum(vals)], total=(style == "total"))
        r += 1


# ============================================================
# Payroll W3
# ============================================================
def sheet_w3(wb):
    ws = wb.create_sheet("Payroll W3")
    write_header(ws, "Payroll Detail (W3)", ["Y1 Gross", "Y2 Gross", "Y3 Gross"], [50, 14, 14, 14])

    roles = [
        ("Founder A — CEO (Yusuf Quadri), £500/mo from Y3 only", 0, 0, 6_000),
        ("Founder B — CTO (Yusuff Adeagbo), £500/mo from Y3 only", 0, 0, 6_000),
        ("Safeguarding & Trust Lead (part-time, ~10 hrs/wk, Y3 only — contingent on 2+ paying universities)", 0, 0, 8_000),
    ]
    r = 5
    totals = [0, 0, 0]
    for label, y1, y2, y3 in roles:
        write_row(ws, r, label, [y1, y2, y3], money=True)
        totals[0] += y1; totals[1] += y2; totals[2] += y3
        r += 1
    write_row(ws, r, "Total gross salaries", totals, total=True); r += 2

    ni = (PL["ni"][0], PL["ni"][1], PL["ni"][2])
    pen = (PL["pension"][0], PL["pension"][1], PL["pension"][2])
    ben = (PL["benefits"][0], PL["benefits"][1], PL["benefits"][2])

    write_row(ws, r, "Employer NI (13.8% above secondary threshold)", list(ni), total=True); r += 1
    write_row(ws, r, "Employer pension (3% above £6,240)", list(pen), total=True); r += 1
    write_row(ws, r, "Other employment costs (kit, training)", list(ben), total=True); r += 1
    write_row(ws, r, "TOTAL EMPLOYMENT COST",
              [totals[i] + ni[i] + pen[i] + ben[i] for i in range(3)], total=True); r += 2

    ws.cell(row=r, column=1,
            value="No team member — founders included — draws any salary in Y1 or Y2. Both founders fund "
                  "personal living costs from existing employment/freelance income throughout. Modest founder "
                  "pay and the plan's only funded hire begin in Y3, funded entirely by revenue — not by any "
                  "external funding round, none of which is assumed in this plan. Further hires beyond Year 3 "
                  "would only be made if revenue growth or a future funding round justifies them.").font = italic_small
    ws.merge_cells(start_row=r, start_column=1, end_row=r, end_column=4)


# ============================================================
# Marketing W4
# ============================================================
def sheet_w4(wb):
    ws = wb.create_sheet("Marketing W4")
    write_header(ws, "Marketing Expenditure Detail (W4)", ["Year 1", "Year 2", "Year 3", "TOTAL"],
                 [44, 14, 14, 14, 14])
    rows = [
        ("Total Marketing Expenses", PL["marketing"], "total"),
        ("", None, False),
        ("Channel-level spend", None, "sub"),
        ("University partnership & PR (NatWest network)", [300, 700, 1_300], False),
        ("Content / SEO (founder-written)", [150, 400, 700], False),
        ("LinkedIn / paid B2B", [0, 300, 700], False),
        ("Instagram / TikTok (organic)", [150, 500, 1_000], False),
        ("Google Search ads (none in Y1)", [0, 200, 500], False),
        ("Student-rep / ambassador programme", [300, 400, 800], False),
        ("", None, False),
        ("KPIs", None, "sub"),
        ("Marketing as % of revenue", ["n/a", "15.6%", "10.4%"], False),
        ("Avg CAC per university (founder-led, low-cost)", ["n/a", "~£2,500", "~£3,300"], False),
    ]
    r = 5
    for label, vals, style in rows:
        if vals is None:
            if style == "sub":
                sub_row(ws, r, label, 4)
            r += 1; continue
        is_money = all(isinstance(v, (int, float)) for v in vals)
        total_val = sum(vals) if is_money else ""
        write_row(ws, r, label, list(vals) + [total_val], total=(style == "total"), money=is_money)
        r += 1


# ============================================================
# R&D W5
# ============================================================
def sheet_w5(wb):
    ws = wb.create_sheet("R&D W5")
    write_header(ws, "Research & Development (W5)", ["Year 1", "Year 2", "Year 3"], [50, 14, 14, 14])
    rows = [
        ("R&D delivery model", None, "sub"),
        ("Founder time (unpaid, no cash cost)", ["Y1-Y2: 100%", "Y1-Y2: 100%", "Some paid capacity from Y3"], False),
        ("Cash R&D spend (already included in COGS/OPEX above)",
         [PL["cogs_llm"][0], PL["cogs_llm"][1], PL["cogs_llm"][2]], True),
        ("", None, False),
        ("SME R&D Tax Credit", None, "sub"),
        ("Assumed as a cash inflow in this model", [0, 0, 0], True),
    ]
    r = 5
    for label, vals, style in rows:
        if vals is None:
            if style == "sub":
                sub_row(ws, r, label, 3)
            r += 1; continue
        write_row(ws, r, label, vals, total=(style is True), money=all(isinstance(v, (int, float)) for v in vals))
        r += 1
    r += 1
    ws.cell(row=r, column=1,
            value="R&D is delivered as unpaid founder time throughout Y1-Y2, plus modest paid capacity from Y3 "
                  "once revenue supports it. No SME R&D Tax Credit inflow is assumed anywhere in this model — "
                  "qualifying PAYE spend is minimal while founders are unpaid. The founders will explore claiming "
                  "R&D tax relief with NatWest's in-kind accountancy support once qualifying costs exist from Y3, "
                  "as a possible upside not relied upon here.").font = italic_small
    ws.merge_cells(start_row=r, start_column=1, end_row=r, end_column=4)


# ============================================================
# Fixed Assets
# ============================================================
def sheet_fa(wb):
    ws = wb.create_sheet("Fixed Assets")
    write_header(ws, "Fixed Asset Schedule",
                 ["Depreciation rate", "Y1 additions", "Y2 additions", "Y3 additions",
                  "Y1 depreciation", "Y2 depreciation", "Y3 depreciation", "Y3 NBV"],
                 [44, 14] + [12] * 7)
    # Split reconciles exactly to DEPR_Y = [-150, -400, -700] used in the P&L.
    rows = [
        ("Tangible — Laptops & equipment", "33%", 500, 500, 900, -130, -340, -580, 850),
        ("Intangible — Capitalised R&D (W5)", "20%", 100, 100, 300, -20, -60, -120, 300),
        ("Total CAPEX", "", 600, 600, 1_200, -150, -400, -700, 1_150),
    ]
    r = 5
    for row in rows:
        label = row[0]; rate = row[1]; vals = row[2:]
        cell = ws.cell(row=r, column=1, value=label)
        cell.font = bold if "Total" in label else regular
        cell.border = border_all; cell.alignment = left_a
        ws.cell(row=r, column=2, value=rate).alignment = right_a
        ws.cell(row=r, column=2).border = border_all
        for j, v in enumerate(vals, start=3):
            c = ws.cell(row=r, column=j, value=v)
            c.alignment = right_a; c.border = border_all
            if isinstance(v, (int, float)):
                c.number_format = gbp(v)
            if "Total" in label:
                c.fill = total_fill; c.font = bold
        r += 1

    ws.cell(row=r + 1, column=1,
            value="MVP is already built so upfront capex is minimal. Y1 hardware = 2 founder laptops; Y3 adds "
                  "a laptop for the first part-time hire. Intangible adds are the cost of capitalised R&D "
                  "meeting IAS 38 criteria.").font = italic_small
    ws.merge_cells(start_row=r + 1, start_column=1, end_row=r + 1, end_column=9)


# ============================================================
# Accelerator Value — quantify in-kind support (Y1 only)
# ============================================================
def sheet_accelerator(wb):
    ws = wb.create_sheet("Accelerator Value")
    write_header(ws, "NatWest Accelerator — Quantified In-Kind Value (Year 1 Only)",
                 ["Annual value (£)", "Cash cost to DEQUAD"], [50, 18, 18])
    rows = [
        ("Office co-working space (3 desks, central London)", 12_000, 0),
        ("Legal advice (Mishcon de Reya, DLA Piper panels)", 4_500, 0),
        ("Accountancy support (PwC alumni network)", 3_600, 0),
        ("Banking & business introductions", 2_000, 0),
        ("Investor pitch coaching & mentoring", 5_000, 0),
        ("Programme demo day & PR placement", 4_000, 0),
        ("Out-of-programme costs (Companies House, IP filings)", 0, 300),
        ("", None, None),
        ("TOTAL in-kind value (Y1)", 31_100, 300),
    ]
    r = 5
    for row in rows:
        label, val, cash = row
        if val is None:
            r += 1; continue
        c = ws.cell(row=r, column=1, value=label); c.border = border_all; c.alignment = left_a
        c.font = bold if label.startswith("TOTAL") else regular
        c2 = ws.cell(row=r, column=2, value=val); c2.alignment = right_a
        c2.border = border_all; c2.number_format = gbp(val)
        c3 = ws.cell(row=r, column=3, value=cash); c3.alignment = right_a
        c3.border = border_all; c3.number_format = gbp(cash)
        if label.startswith("TOTAL"):
            for cc in (c, c2, c3): cc.fill = total_fill; cc.font = bold
        r += 1
    ws.cell(row=r + 1, column=1,
            value="The NatWest Accelerator membership materially de-risks Year 1: £31k of services received "
                  "in-kind against a few hundred pounds of cash cost, covered within the £6,000 founder capital. "
                  "This support runs for Year 1 only — from Year 2 the plan budgets modest real cash costs for "
                  "office, legal and accountancy (see Cash Flow 3yr).").font = italic_small
    ws.merge_cells(start_row=r + 1, start_column=1, end_row=r + 1, end_column=4)


# ============================================================
# Startup loan placeholder
# ============================================================
def sheet_loan(wb):
    ws = wb.create_sheet("Startup Loan")
    write_header(ws, "Startup Loan Schedule",
                 ["Principal", "Interest rate", "Term (months)", "Monthly payment"], [50] + [14] * 4)
    ws.cell(row=5, column=1, value="No external debt or equity taken at any point in this 3-year plan.").font = bold
    ws.cell(row=6, column=1,
            value="The founders are self-funding the entire 3-year plan with £6,000 of equity plus revenue, "
                  "and the in-kind NatWest Accelerator support in Year 1. No pre-seed, seed or other external "
                  "investment is assumed, committed, or required for this plan to succeed. If institutional "
                  "traction significantly exceeds this conservative forecast, the founders may explore external "
                  "investment beyond Year 3 — that scenario is outside the scope of this model.").alignment = left_a
    ws.merge_cells(start_row=6, start_column=1, end_row=6, end_column=5)


# ============================================================
# Build the workbook
# ============================================================
def main():
    wb = Workbook()
    sheet_readme(wb)
    sheet_pl(wb)
    sheet_cf(wb)
    sheet_cf_monthly(wb)
    sheet_bs(wb)
    sheet_w1(wb)
    sheet_w2(wb)
    sheet_w3(wb)
    sheet_w4(wb)
    sheet_w5(wb)
    sheet_fa(wb)
    sheet_accelerator(wb)
    sheet_loan(wb)
    wb.save(OUT)
    print(f"Wrote {OUT}")
    print(f"  Y1 revenue: £{REV_Y[0]:,}  Y2: £{REV_Y[1]:,}  Y3: £{REV_Y[2]:,}")
    print(f"  Y1 opening cash: £{CASH_OPEN[0]:,}  Y1 closing cash: £{CASH_CLOSE[0]:,}")
    print(f"  Y3 closing cash: £{CASH_CLOSE[2]:,}")


if __name__ == "__main__":
    main()
